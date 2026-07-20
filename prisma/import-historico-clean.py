# -*- coding: utf-8 -*-
import sys, io, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

# Limpia la matriz histórica de la hoja "PADRON 2022" a un JSON que consume
# prisma/import-historico.ts.
#   python prisma/import-historico-clean.py "<ruta al .xlsx>" prisma/_historico.json
#
# La hoja NO es un padrón: es una matriz indexada por PUESTO con un titular por
# empadronamiento. Cada fila = un puesto; columnas F..N = las cuatro gestiones.
XLSX = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\anonimo\Documents\2026\mercado milagros\2026 ARCHIVOS\PADRON\PADRON EN EXCEL\RELACION DE PUESTOS (version 5).xlsx"
OUT  = sys.argv[2] if len(sys.argv) > 2 else "prisma/_historico.json"

# read_only=False es OBLIGATORIO: en modo read_only openpyxl no expone
# `merged_cells`, y etapa/bloque vienen combinados verticalmente. Sin eso, las
# 704 filas quedan sin etapa.
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["PADRON 2022"]

# Propaga el valor de la celda superior-izquierda a todo el rango combinado.
merged = {}
for rng in ws.merged_cells.ranges:
    top = ws.cell(rng.min_row, rng.min_col).value
    for r in range(rng.min_row, rng.max_row + 1):
        for c in range(rng.min_col, rng.max_col + 1):
            merged[(r, c)] = top

def val(r, c):
    v = merged.get((r, c), ws.cell(r, c).value)
    return v.strip() if isinstance(v, str) else v

def txt(r, c):
    v = val(r, c)
    if v is None:
        return None
    s = str(v).strip()
    return s or None

def num(r, c):
    v = val(r, c)
    if v is None:
        return None
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else None

COLS = {"etapa": 1, "bloque": 2, "numero": 4,
        "n2014": 6, "p2014": 7, "n2017": 8, "p2017": 9,
        "n2019": 10, "p2019": 11, "n2021": 12, "dni": 13, "p2021": 14}

out, descartadas = [], []
for r in range(4, ws.max_row + 1):
    numero = num(r, COLS["numero"])
    if numero is None:
        continue
    et_raw = val(r, COLS["etapa"])
    bloque = txt(r, COLS["bloque"])
    # OJO: la celda de etapa de la Etapa 1 vale 0, que es FALSY en Python. Un
    # `if not et_raw` descarta en silencio las 296 filas de E1. Comparar con None.
    if et_raw is None or bloque is None:
        # Pie de tabla del Excel ("PUESTOS SIN EMPADRONAR", 296, 408): totales.
        descartadas.append((r, txt(r, COLS["n2021"]), numero))
        continue
    etapa = 2 if "SEGUNDA" in str(et_raw).upper() else 1
    dni = val(r, COLS["dni"])
    out.append({
        "filaExcel": r, "etapa": etapa, "bloque": bloque.upper(), "numero": numero,
        "e2014": {"nombre": txt(r, COLS["n2014"]), "padron": num(r, COLS["p2014"])},
        "e2017": {"nombre": txt(r, COLS["n2017"]), "padron": num(r, COLS["p2017"])},
        "e2019": {"nombre": txt(r, COLS["n2019"]), "padron": num(r, COLS["p2019"])},
        "e2021": {"nombre": txt(r, COLS["n2021"]), "padron": num(r, COLS["p2021"]),
                  "dni": None if dni is None else str(dni).strip()},
    })

claves = [f"E{o['etapa']}-{o['bloque']}-{o['numero']}" for o in out]
assert len(claves) == len(set(claves)), "claves de puesto duplicadas en el Excel"

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"filas emitidas : {len(out)}")
print(f"descartadas    : {len(descartadas)} {descartadas}")
for k in ["e2014", "e2017", "e2019", "e2021"]:
    print(f"  {k}: titulares={sum(1 for o in out if o[k]['nombre'])}"
          f" padron={sum(1 for o in out if o[k]['padron'] is not None)}")
print(f"→ {OUT}")
